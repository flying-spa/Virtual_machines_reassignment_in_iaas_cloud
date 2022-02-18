'''from astar import *'''
import xlrd
import xlwt
import ss
from openpyxl import load_workbook

book=xlrd.open_workbook('C:/Users/sutha/AppData/Local/Programs/Python/Python37-32/Mini Project/data.xlsx')
#book=xlrd.open_workbook('C:/Users/sutha/AppData/Local/Programs/Python/Python37-32/Mini Project/data_1.xlsx')
sheet=book.sheet_by_name('Sheet2')
datag=[[sheet.cell_value(r,c) for c in range(sheet.ncols)]for r in range(sheet.nrows)]
#goal state
#initial state
sheet1=book.sheet_by_name('Sheet3')
datai=[[sheet1.cell_value(r,c) for c in range(sheet1.ncols)]for r in range(sheet1.nrows)]
goal_state=[]
initial_state=[]
for i in range(len(datai)):
    temp=[]
    for j in range(len(datai[i])):
        if datai[i][j]!='':
            temp.append(int(datai[i][j]))
    if(sum(temp)!=0):
        initial_state.append(temp)
for i in range(len(initial_state)):
    temp=[]
    for j in range(len(datag[i])):
        temp.append(int(datag[i][j]))
    goal_state.append(temp)

#print("Goal State\n",goal_state)
x=len(goal_state)
#print(type(x),"x :",x)
y=len(goal_state[x-1])

sheet3 = book.sheet_by_name('Sheet1')
rows = []
for i in range(sheet3.ncols):
    columns = []
    for j in range(sheet3.nrows):
        x1=sheet3.cell(j, i).value
        #print(x)
        if x1!='':
            columns.append(x1)
    if len(columns)!=0:
        rows.append(columns)
#print (rows)
        
dem_ram1=[]
dem_cpu1=[]
min_clust=0
for i in range(len(rows)):
    if rows[i][0]=='Dem_ram':
        for j in range(1,21):
            dem_ram1.append(int(rows[i][j]))
    elif rows[i][0]=='Dem_cpu':
        for j in range(1,21):
            dem_cpu1.append(int(rows[i][j]))
    elif rows[i][0]=='min_clusters':
        for j in range(1,2):
            min_clust=(int(rows[i][j]))
#print("dem ram: ",dem_ram1,"len of dem ram: ",len(dem_ram1),"data type: ",type(dem_ram1))
print("dem cpu: ",dem_cpu1,"len of dem cap: ",len(dem_cpu1),"data type: ",type(dem_cpu1))

book1=xlrd.open_workbook('C:/Users/sutha/AppData/Local/Programs/Python/Python37-32/Mini Project/trial.xlsx')
sheet4=book1.sheet_by_name('Sheet 1')
rows1 = []
for i in range(sheet4.ncols):
    columns1 = []
    for j in range(sheet4.nrows):
        x2=sheet4.cell(j, i).value
        #print(x)
        if x2!='':
            columns1.append(x2)
    if len(columns1)!=0:
        rows1.append(columns1)
#print (rows1)

        
if __name__=="__main__":
    
    #print("No.of rows1: ",len(goal_state),"\nNo.of columns1: ",len(goal_state[x-1]))
    #print("Initail State\n",initial_state)
    print("Goal state:\n",goal_state,"\ninitial state: \n",initial_state)
    print("\nAfter Executing astar Algorithm\n")
    exec(open("astar.py").read())
    print("flag_astar: ",ss.flag_astar)
    print("time_astar: ",ss.time_astar)
    print("\nAfter Executing dmh Algorithm\n")
    exec(open("dmh.py").read())
    print("flag_dmh: ",ss.flag_dmh)
    print("time_dmh: ",ss.time_dmh)
    print("\nAfter Executing idmh Algorithm\n")
    exec(open("idmh.py").read())
    print("flag_idmh: ",ss.flag_idmh)
    print("time_idmh: ",ss.time_idmh)


    workbook_name = 'C:/Users/sutha/AppData/Local/Programs/Python/Python37-32/Mini Project/trial.xlsx'
    wb = load_workbook(workbook_name)
    page = wb.active

    # New data to write:
    new_companies = [len(goal_state[0]),len(initial_state),min_clust,ss.flag_astar,ss.time_astar,ss.space_astar,ss.moves_astar,ss.flag_dmh,ss.time_dmh,ss.moves_dmh,ss.flag_idmh,ss.time_idmh,ss.moves_idmh]
    page.append(new_companies)

    wb.save(filename=workbook_name)

